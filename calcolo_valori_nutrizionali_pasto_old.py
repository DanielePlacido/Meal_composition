# Placido Daniele
# 29/06/2020
# This script evaluate the amount of CHO, proteins, lipdes, fiber, water and \
# the energy of a meal, plots the proportions of these constituents and \
# evaluate the units of insulin to do to metabolize carbohydrates, together \
# with the fraction of fast and slow absorptions CHO. An warning message is \
# printed if the total weight of CHO is below 80% or above 120% of the \
# recommended value per meal.

# If I have time I should add the evaluation of the recommended fraction of constituents per meal and plot it

from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import os.path as osp
import warnings
import matplotlib.pyplot as plt
from matplotlib.patches import ConnectionPatch

import tkinter as tk
from tkinter import messagebox

# -- Script starts here --

## DATA

cho_meal = 120. # g recommended cho weight per meal
protein_meal = 44. # g recommended proteins weight per meal
lipid_meal = 44. # g recommended lipides weight per meal

#energy

def Selection(meal):

	"""
	Function tha shows the selected meal
	"""

	selection = "You selected the option " + meal
	frm_selection = tk.Frame(master = frm_main, relief = tk.SUNKEN, \
									borderwidth = 3)
	lbl_selection = tk.Label(master = frm_selection, text = selection)

	frm_selection.pack(fill = tk.X)
	lbl_selection.pack()
	btn_continue["state"] = "normal"

def Evaluate_meal(meal):

	"""
	Function that quits the main loop allowing user to go on with evaluation of meal composition and units to be injected.
	"""

	window.withdraw()

	if meal == "breakfast":
		RR = 10.  # insulin cho ratio
		meal_up = "BREAKFAST"
	elif meal == "lunch":
		RR = 9. # insulin cho ratio
		meal_up = "LUNCH"
	elif meal == "dinner":
		RR = 10. # insulin cho ratio
		meal_up = "DINNER"
	# open data base
	folder_path = osp.realpath("")
	# check on path: prints True if path exists, False otherwise
	#print(osp.exists(folder_path))
	file_path = osp.join(folder_path,"Tabelle_valori_nutrizionali_alimenti.xlsx")
	#print(osp.exists(file_path))
	data_base = load_workbook(file_path,data_only=True)

	sheet_list = data_base.get_sheet_names()
	#print(sheet_list)

	weight_tot = 0.0
	cho_tot = 0.0
	cho_fast_tot = 0.0
	cho_slow_tot = 0.0
	protein_tot = 0.0
	lipid_tot = 0.0
	fiber_tot = 0.0
	water_tot = 0.0
	energy_tot = 0.0

	# read data base
	for sheet_name in sheet_list:
		sheet = data_base[sheet_name]
		flag_read = int(sheet.cell(row = 1, column = 2).value)
		if flag_read == 1: # read sheet content
			r_min = sheet.min_row + 2
			r_max = sheet.max_row
			c_min = sheet.min_column
			c_max = sheet.max_column

			weight = 0.0
			cho = 0.0
			cho_fast = 0.0
			cho_slow = 0.0
			protein = 0.0
			lipid = 0.0
			fiber = 0.0
			water = 0.0
			energy = 0.0

			for peso in sheet.iter_cols(min_row = r_min, max_row = r_max,\
																		min_col = c_min + 1, max_col = c_min + 1,\
																		values_only = True):
				peso_alimento = np.array(peso)

			ind_peso = np.nonzero(peso_alimento > 0)
			peso_alimento = peso_alimento[ind_peso[0]]
			weight = np.sum(peso_alimento) # g ingredients weight

			MM = np.zeros((ind_peso[0].shape[0], c_max - 1))
			for rr in range(len(ind_peso[0])):
				for cc in range(c_min + 1, c_max + 1):
					MM[rr,cc-2] = float(sheet.cell(row = ind_peso[0][rr] + 3, column = cc).value)

			if sheet_name == "Verdure":
				cho = np.sum(MM[:,0]*MM[:,1]*.8) # evaluate cho introducing a correction coefficient to keep into account of fiber effects, weight in g
			else:
				cho = np.sum(MM[:,0]*MM[:,1]) # g cho weight
			protein = np.sum(MM[:,0]*MM[:,2]) # g protein weight
			lipid = np.sum(MM[:,0]*MM[:,3]) # g lipid weight
			fiber = np.sum(MM[:,0]*MM[:,4]) # g fiber weight
			water = weight - (cho + protein + lipid + fiber) # g water weight
			energy = np.sum(MM[:,0]*MM[:,5]/100) # kcal energy

			if sheet_name == "Verdure" or sheet_name == "Frutta" or \
				 sheet_name == "Dolci" or sheet_name == "Bevande" or \
				 sheet_name == "Aperitivi":
				# this are fast cho
				cho_fast = cho
			elif sheet_name == "Pasta riso e cereali" or \
					 sheet_name == "Prodotti da forno e cereali" or \
					 sheet_name == "Legumi" or sheet_name == "Stuzzichini":
				# this are slow cho
				cho_slow = cho

			weight_tot = weight_tot + weight # g total meal weight evaluation
			water_tot = water_tot + water # g total water weight
			cho_tot = cho_tot + cho # g total cho weight
			cho_fast_tot = cho_fast_tot + cho_fast # g total fast cho weight
			cho_slow_tot = cho_slow_tot + cho_slow # g total slow cho weight
			protein_tot = protein_tot + protein # g total protein weight
			lipid_tot = lipid_tot + lipid # g total lipid weight
			fiber_tot = fiber_tot + fiber # g total fiber weight
			energy_tot = energy_tot + energy # kcal total meal energy

	weight_dry = weight_tot - water_tot # g total dry meal weight evaluation
	cho_frac = cho_tot/weight_dry
	protein_frac = protein_tot/weight_dry
	lipid_frac = lipid_tot/weight_dry
	fiber_frac = fiber_tot/weight_dry
	# water_frac = water_tot/weight_dry

	cho_fast_frac = cho_fast_tot/cho_tot
	cho_slow_frac = 1.0 - cho_fast_frac

	unit = round(cho_tot/RR,1) # number of insulin units to inject to metabolize 	meal
	
	# text for the GUI dialog box
	message_text = f"** {meal_up} SUMMARY **\nTotal {meal} weight: {weight_tot} g\nTotal {meal} weight dry: {round(weight_dry)} g\nCHO from this {meal}: {round(cho_tot)} g\nProteins from this {meal}: {round(protein_tot)} g\nLipides from this {meal}: {round(lipid_tot)} g\nFiber from this {meal}: {round(fiber_tot)} g\nEnergy from this {meal}: {round(energy_tot)} kcal\nUnits to inject: {unit}\n"
	# create dialog box
	messagebox.showinfo(message = message_text)
	## OUTPUT
	print(f"\n** {meal_up} SUMMARY **\n")
	print(f"Total {meal} weight: {weight_tot} g\n")
	print(f"Total {meal} weight dry: {round(weight_dry)} g\n")
	print(f"CHO from this {meal}: {round(cho_tot)} g\n")
	print(f"Proteins from this {meal}: {round(protein_tot)} g\n")
	print(f"Lipides from this {meal}: {round(lipid_tot)} g\n")
	print(f"Fiber from this {meal}: {round(fiber_tot)} g\n")
	print(f"Units to inject: {unit}\n")
	print(f"Energy from this {meal}: {round(energy_tot)} kcal\n")

	# warnings about cho g in the meal
	if round(cho_tot) < cho_meal*.8:
		warning_text = f"\nYou are below the cho threshold for this meal:\n{round(cho_tot)} < {cho_meal*.8}\nRemember that you should heat {cho_meal} g of CHO!\n"
		warnings.warn(warning_text)
		messagebox.showwarning(message = warning_text)
	elif round(cho_tot) > cho_meal*1.2:
		warning_text = f"\nYou are above the cho threshold for this meal:\n{round(cho_tot)} > {cho_meal*1.2}\nRemember that you should heat {cho_meal} g of CHO!\n"
		warnings.warn(warning_text)
		messagebox.showwarning(message = warning_text)
	# warnings about proteins g in the meal
	if round(protein_tot) < protein_meal*.8:
		warning_text = f"\nYou are below the proteins threshold for this meal:\n{round(protein_tot)} < {round(protein_meal*.8,1)}\nRemember that you should heat {protein_meal} g of proteins!\n"
		warnings.warn(warning_text)
		messagebox.showwarning(message = warning_text)
	elif round(protein_tot) > protein_meal*1.1:
		warning_text = f"\nYou are above the proteins threshold for this meal:\n{round(protein_tot)} > {round(protein_meal*1.1,1)}\nRemember that you should heat {protein_meal} g of proteins!\n"
		warnings.warn(warning_text)
		messagebox.showwarning(message = warning_text)
	# warnings about lipides g in the meal
	if round(lipid_tot) < lipid_meal*.8:
		warning_text = f"\nYou are below the lipides threshold for this meal:\n{round(lipid_tot)} < {round(lipid_meal*.8,1)}\nRemember that you should heat {lipid_meal} g of lipides!\n"
		warnings.warn(warning_text)
		messagebox.showwarning(message = warning_text)
	elif round(lipid_tot) > lipid_meal*1.1:
		warning_text = f"\nYou are above the lipides threshold for this meal:\n{round(lipid_tot)} > {round(lipid_meal*1.1,1)}\nRemember that you should heat {lipid_meal} g of lipides!\n"
		warnings.warn(warning_text)
		messagebox.showwarning(message = warning_text)

	# HERE STARTS THE FOOD DIARY COMPILATION
	file_name = "food_diary.xlsx"
	diary_path = osp.join(folder_path,file_name)
	print(osp.exists(diary_path))
	if osp.exists(diary_path) == False:
		# Create workbook food_diary.xlsx
		workbook = Workbook()
		sheet = workbook.active
		workbook.save(filename = file_name)
	else:
		# just go haed and compile the diary
		print(f"Workbook {file_name} already exists\n")

	# PLOT
	# make figure and assign axis objects
	fig = plt.figure(num = f"{meal} composition [%]" ,figsize=(9, 5))

	ax1 = fig.add_subplot(121)
	ax2 = fig.add_subplot(122)

	fig.subplots_adjust(wspace=0)

	# Pie chart:
	label = "CHO", "Proteins","Lipides","Fibers"
	sizes = [cho_frac, protein_frac, lipid_frac, fiber_frac]
	explode = (0.1, 0, 0, 0) # only "explode" the 1st slice (i.e. "CHO")
	# rotate so that first wedge is split by the x-axis
	angle = -180*sizes[0]
	ax1.pie(sizes, explode=explode, labels=label, autopct='%1.1f%%', 	startangle=angle ,shadow=False)
	ax1.axis('equal') # Equal aspect ratio ensures that pie is drawn as a circle.

	# bar chart parameters

	xpos = 0
	bottom = 0
	ratios = [cho_fast_frac, cho_slow_frac]
	width = .2
	colors = ["c", "g"]

	for j in range(len(ratios)):
		height = ratios[j]
		ax2.bar(xpos, height, width, bottom=bottom, color=colors[j])
		ypos = bottom + ax2.patches[j].get_height()/2
		bottom += height
		ax2.text(xpos, ypos, "%d%%" % (ax2.patches[j].get_height()*100),ha='center')

	ax2.set_title('CHO composition')
	ax2.legend(("fast","slow"))
	ax2.axis('off')
	ax2.set_xlim(- 2.5*width, 2.5*width)

	# use ConnectionPatch to draw lines between the two plots
	# get the wedge data
	theta1, theta2 = ax1.patches[0].theta1, ax1.patches[0].theta2
	center, r = ax1.patches[0].center, ax1.patches[0].r
	bar_height = sum([item.get_height() for item in ax2.patches])

	# draw top connecting line
	x = r*np.cos(np.pi/180*theta2) + center[0]
	y = np.sin(np.pi/180*theta2) + center[1]
	con = ConnectionPatch(xyA=(-width/2, bar_height), coordsA=ax2.transData,
												xyB=(x, y), coordsB=ax1.transData)
	con.set_color([0, 0, 0])
	con.set_linewidth(4)
	ax2.add_artist(con)

	# draw bottom connecting line
	x = r*np.cos(np.pi/180*theta1) + center[0]
	y = np.sin(np.pi/180*theta1) + center[1]
	con = ConnectionPatch(xyA=(-width/2,0), coordsA=ax2.transData,
												xyB=(x, y), coordsB=ax1.transData)
	con.set_color([0, 0, 0])
	ax2.add_artist(con)
	con.set_linewidth(4)

	plt.show()

	window.quit()

# end function Continue

window = tk.Tk()

frm_main = tk.Frame(master = window, relief = tk.SUNKEN, borderwidth = 3)
frm_message = tk.Frame(master = frm_main, relief = tk.SUNKEN, borderwidth = 3)
frm_radio = tk.Frame(master = frm_main, relief = tk.SUNKEN, borderwidth = 3)
frm_continue = tk.Frame(master = frm_main, relief = tk.SUNKEN, borderwidth = 3)

lbl_message = tk.Label(master = frm_message, text = "Select the meal to use the correct Insulin carbohydrates ratio.")

meal = tk.StringVar(value = "no meal")
breakfast = tk.Radiobutton(master = frm_radio, text = "Breakfast", \
						variable = meal, value = "breakfast", \
						command = lambda:Selection(meal.get()))
lunch = tk.Radiobutton(master = frm_radio, text = "Lunch", variable = meal, \
				value = "lunch", command = lambda:Selection(meal.get()))
dinner = tk.Radiobutton(master = frm_radio, text = "Dinner", variable = meal, \
				value = "dinner", command = lambda:Selection(meal.get()))

btn_continue = tk.Button(master = frm_continue, text = "Evaluate meal", \
							 command = lambda: Evaluate_meal(meal.get()), state = "disabled")

frm_main.pack(fill = tk.X)
frm_message.pack(fill = tk.X)
frm_radio.pack(fill = tk.X)
frm_continue.pack(fill = tk.X)

lbl_message.pack()
btn_continue.pack(anchor = "e")
breakfast.pack(anchor = "w")
lunch.pack(anchor = "w")
dinner.pack(anchor = "w")

window.mainloop()
